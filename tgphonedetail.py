import asyncio
import json
import logging
import os
import pickle
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime
import re
from rich.console import Console
from rich.prompt import Prompt, Confirm
from rich import print as rprint
from telethon.sync import TelegramClient, errors
from telethon.tl import types
from telethon.tl.functions.contacts import ImportContactsRequest, DeleteContactsRequest
from telethon.tl.functions.users import GetFullUserRequest
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", handlers=[logging.FileHandler("telegram_checker.log"), logging.StreamHandler()])
logger = logging.getLogger(__name__)
console = Console()
CONFIG_FILE = Path("config.pkl")
RESULTS_DIR = Path("results")

@dataclass
class TelegramUser:
    id: int
    username: Optional[str]
    first_name: Optional[str]
    last_name: Optional[str]
    phone: str
    premium: bool
    verified: bool
    fake: bool
    bot: bool
    last_seen: str
    last_seen_exact: Optional[str] = None
    status_type: Optional[str] = None
    bio: Optional[str] = None
    common_chats_count: Optional[int] = None
    blocked: Optional[bool] = None
    privacy_restricted: bool = False

    @classmethod
    async def from_user(cls, client: TelegramClient, user: types.User, phone: str = "") -> 'TelegramUser':
        try:
            bio = ''
            common_chats_count = 0
            blocked = False
            
            try:
                full_user = await client(GetFullUserRequest(user.id))
                user_full_info = full_user.full_user
                bio = getattr(user_full_info, 'about', '') or ''
                common_chats_count = getattr(user_full_info, 'common_chats_count', 0)
                blocked = getattr(user_full_info, 'blocked', False)
            except:
                pass

            status_info = get_enhanced_user_status(user.status)
            
            return cls(
                id=user.id,
                username=user.username,
                first_name=getattr(user, 'first_name', None) or "",
                last_name=getattr(user, 'last_name', None) or "",
                phone=phone,
                premium=getattr(user, 'premium', False),
                verified=getattr(user, 'verified', False),
                fake=getattr(user, 'fake', False),
                bot=getattr(user, 'bot', False),
                last_seen=status_info['display_text'],
                last_seen_exact=status_info['exact_time'],
                status_type=status_info['status_type'],
                bio=bio,
                common_chats_count=common_chats_count,
                blocked=blocked,
                privacy_restricted=status_info['privacy_restricted']
            )
        except Exception as e:
            logger.error(f"Error creating TelegramUser: {str(e)}")
            status_info = get_enhanced_user_status(getattr(user, 'status', None))
            return cls(
                id=user.id,
                username=getattr(user, 'username', None),
                first_name=getattr(user, 'first_name', None) or "",
                last_name=getattr(user, 'last_name', None) or "",
                phone=phone,
                premium=getattr(user, 'premium', False),
                verified=getattr(user, 'verified', False),
                fake=getattr(user, 'fake', False),
                bot=getattr(user, 'bot', False),
                last_seen=status_info['display_text'],
                last_seen_exact=status_info['exact_time'],
                status_type=status_info['status_type'],
                privacy_restricted=status_info['privacy_restricted']
            )

def get_enhanced_user_status(status: types.TypeUserStatus) -> Dict[str, Any]:
    result = {
        'display_text': 'Unknown',
        'exact_time': None,
        'status_type': 'unknown',
        'privacy_restricted': False
    }
    
    if isinstance(status, types.UserStatusOnline):
        result.update({
            'display_text': "Currently online",
            'status_type': 'online',
            'privacy_restricted': False
        })
    elif isinstance(status, types.UserStatusOffline):
        exact_time = status.was_online.strftime('%Y-%m-%d %H:%M:%S UTC')
        result.update({
            'display_text': f"Last seen: {exact_time}",
            'exact_time': exact_time,
            'status_type': 'offline',
            'privacy_restricted': False
        })
    elif isinstance(status, types.UserStatusRecently):
        result.update({
            'display_text': "Last seen recently (1 second - 3 days ago)",
            'status_type': 'recently',
            'privacy_restricted': True
        })
    elif isinstance(status, types.UserStatusLastWeek):
        result.update({
            'display_text': "Last seen within a week (3-7 days ago)",
            'status_type': 'last_week',
            'privacy_restricted': True
        })
    elif isinstance(status, types.UserStatusLastMonth):
        result.update({
            'display_text': "Last seen within a month (7-30 days ago)",
            'status_type': 'last_month',
            'privacy_restricted': True
        })
    elif status is None:
        result.update({
            'display_text': "Status unavailable",
            'status_type': 'unavailable'
        })
    
    return result

def validate_phone_number(phone: str) -> str:
    phone = re.sub(r'[^\d+]', '', str(phone).strip())
    if not phone.startswith('+'): phone = '+' + phone
    if not re.match(r'^\+\d{10,15}$', phone): raise ValueError(f"Invalid phone number format: {phone}")
    return phone

def validate_username(username: str) -> str:
    username = username.strip().lstrip('@')
    if not re.match(r'^[A-Za-z]\w{3,30}[A-Za-z0-9]$', username): raise ValueError(f"Invalid username format: {username}")
    return username

def read_excel_phones(file_path: str, phone_column: str = "phone") -> List[str]:
    """Read phone numbers from Excel file."""
    try:
        # Try to read Excel file
        if file_path.endswith('.xlsx'):
            df = pd.read_excel(file_path, engine='openpyxl')
        elif file_path.endswith('.xls'):
            df = pd.read_excel(file_path, engine='xlrd')
        else:
            raise ValueError("Unsupported file format. Use .xlsx or .xls")
        
        # Check if the specified column exists
        if phone_column not in df.columns:
            available_columns = list(df.columns)
            console.print(f"[yellow]Column '{phone_column}' not found. Available columns: {available_columns}[/yellow]")
            phone_column = Prompt.ask(f"Enter column name containing phone numbers", choices=available_columns)
        
        # Extract phone numbers and clean them
        phones = df[phone_column].dropna().astype(str).tolist()
        return phones
    
    except Exception as e:
        console.print(f"[red]Error reading Excel file: {str(e)}[/red]")
        # Fallback to reading as text file
        console.print("[yellow]Trying to read as text file...[/yellow]")
        try:
            with open(file_path, 'r') as f:
                phones = [line.strip() for line in f if line.strip()]
            return phones
        except:
            raise ValueError(f"Could not read file: {file_path}")

def save_to_excel(results: dict, filename: str):
    """Save results to Excel file with formatting."""
    data = []
    
    for identifier, result in results.items():
        if "error" in result:
            row = {
                'Phone': identifier,
                'Status': 'Error',
                'Error': result['error'],
                'Username': '',
                'First Name': '',
                'Last Name': '',
                'Full Name': '',
                'Premium': '',
                'Verified': '',
                'Bot': '',
                'Last Seen': '',
                'Last Seen Exact': '',
                'Bio': '',
                'Common Chats': '',
                'Blocked': '',
                'Fake': '',
                'Privacy Restricted': ''
            }
        else:
            full_name = f"{result.get('first_name', '')} {result.get('last_name', '')}".strip()
            row = {
                'Phone': identifier,
                'Status': 'Found',
                'Error': '',
                'Username': f"@{result.get('username', '')}" if result.get('username') else '',
                'First Name': result.get('first_name', ''),
                'Last Name': result.get('last_name', ''),
                'Full Name': full_name,
                'Premium': 'Yes' if result.get('premium') else 'No',
                'Verified': 'Yes' if result.get('verified') else 'No',
                'Bot': 'Yes' if result.get('bot') else 'No',
                'Last Seen': result.get('last_seen', ''),
                'Last Seen Exact': result.get('last_seen_exact', ''),
                'Bio': result.get('bio', ''),
                'Common Chats': result.get('common_chats_count', 0),
                'Blocked': 'Yes' if result.get('blocked') else 'No',
                'Fake': 'Yes' if result.get('fake') else 'No',
                'Privacy Restricted': 'Yes' if result.get('privacy_restricted') else 'No'
            }
        data.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to Excel with formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        
        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Set column widths
        column_widths = {
            'A': 20, 'B': 10, 'C': 30, 'D': 20, 'E': 15, 
            'F': 15, 'G': 20, 'H': 10, 'I': 10, 'J': 10,
            'K': 25, 'L': 20, 'M': 40, 'N': 12, 'O': 10,
            'P': 10, 'Q': 18
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Apply formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format data rows
        for row in range(2, len(df) + 2):
            status_cell = worksheet[f'B{row}']
            if status_cell.value == 'Found':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            elif status_cell.value == 'Error':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")
    
    console.print(f"[green]Results saved to Excel file: {filename}[/green]")

class TelegramChecker:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        RESULTS_DIR.mkdir(exist_ok=True)

    def load_config(self) -> dict:
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, 'rb') as f: return pickle.load(f)
            except Exception as e:
                logger.error(f"Error loading config: {e}")
                return {}
        return {}

    def save_config(self):
        with open(CONFIG_FILE, 'wb') as f: pickle.dump(self.config, f)

    async def initialize(self):
        if not self.config.get('api_id'):
            console.print("[yellow]First time setup - please enter your Telegram API credentials[/yellow]")
            console.print("[cyan]You can get these from https://my.telegram.org/apps[/cyan]")
            self.config['api_id'] = 37597265
            self.config['api_hash'] = "650a8b45cb705150a2d3bb7f6cd41bee"
            self.config['phone'] = "+8801635031519"  # Added + for phone number
            self.save_config()

        self.client = TelegramClient('telegram_checker_session', self.config['api_id'], self.config['api_hash'])
        await self.client.connect()
        
        if not await self.client.is_user_authorized():
            await self.client.send_code_request(self.config['phone'])
            code = Prompt.ask("Enter the verification code sent to your Telegram")
            try:
                await self.client.sign_in(self.config['phone'], code)
            except errors.SessionPasswordNeededError:
                password = Prompt.ask("Enter your 2FA password", password=True)
                await self.client.sign_in(password=password)

    async def check_phone_number(self, phone: str) -> Optional[TelegramUser]:
        try:
            phone = validate_phone_number(phone)
            try:
                user = await self.client.get_entity(phone)
                telegram_user = await TelegramUser.from_user(self.client, user, phone)
                return telegram_user
            except:
                contact = types.InputPhoneContact(client_id=0, phone=phone, first_name="Test", last_name="User")
                result = await self.client(ImportContactsRequest([contact]))
                
                if not result.users: return None
                
                user = result.users[0]
                try:
                    full_user = await self.client.get_entity(user.id)
                    await self.client(DeleteContactsRequest(id=[user.id]))
                    telegram_user = await TelegramUser.from_user(self.client, full_user, phone)
                    return telegram_user
                finally:
                    try:
                        await self.client(DeleteContactsRequest(id=[user.id]))
                    except:
                        pass
        except Exception as e:
            logger.error(f"Error checking {phone}: {str(e)}")
            return None

    async def check_username(self, username: str) -> Optional[TelegramUser]:
        try:
            username = validate_username(username)
            user = await self.client.get_entity(username)
            if not isinstance(user, types.User): return None
            telegram_user = await TelegramUser.from_user(self.client, user, "")
            return telegram_user
        except ValueError as e:
            logger.error(f"Invalid username {username}: {str(e)}")
            return None
        except errors.UsernameNotOccupiedError:
            logger.error(f"Username {username} not found")
            return None
        except Exception as e:
            logger.error(f"Error checking username {username}: {str(e)}")
            return None

    async def process_phones(self, phones: List[str]) -> dict:
        results = {}
        total_phones = len(phones)
        console.print(f"\n[cyan]Processing {total_phones} phone numbers...[/cyan]")
        
        for i, phone in enumerate(phones, 1):
            try:
                phone = phone.strip()
                if not phone: continue
                console.print(f"[cyan]Checking {phone} ({i}/{total_phones})[/cyan]")
                user = await self.check_phone_number(phone)
                results[phone] = asdict(user) if user else {"error": "No Telegram account found"}
            except ValueError as e:
                results[phone] = {"error": str(e)}
            except Exception as e:
                results[phone] = {"error": f"Unexpected error: {str(e)}"}
        return results

    async def process_usernames(self, usernames: List[str]) -> dict:
        results = {}
        total_usernames = len(usernames)
        console.print(f"\n[cyan]Processing {total_usernames} usernames...[/cyan]")
        
        for i, username in enumerate(usernames, 1):
            try:
                username = username.strip()
                if not username: continue
                console.print(f"[cyan]Checking {username} ({i}/{total_usernames})[/cyan]")
                user = await self.check_username(username)
                results[username] = asdict(user) if user else {"error": "No Telegram account found"}
            except ValueError as e:
                results[username] = {"error": str(e)}
            except Exception as e:
                results[username] = {"error": f"Unexpected error: {str(e)}"}
        return results

def display_summary(results: dict):
    """Display summary of results."""
    total = len(results)
    found = sum(1 for r in results.values() if "error" not in r)
    errors = total - found
    
    console.print("\n" + "="*50)
    console.print("[bold]RESULTS SUMMARY[/bold]")
    console.print("="*50)
    console.print(f"[green]✓ Accounts Found: {found}[/green]")
    console.print(f"[red]✗ Errors/Not Found: {errors}[/red]")
    console.print(f"[cyan]Total Processed: {total}[/cyan]")
    console.print("="*50)
    
    if found > 0:
        console.print("\n[bold]FOUND ACCOUNTS:[/bold]")
        for identifier, data in results.items():
            if "error" not in data:
                name = f"{data.get('first_name', '')} {data.get('last_name', '')}".strip()
                username = f"@{data.get('username', '')}" if data.get('username') else 'No username'
                console.print(f"  • {identifier}: {name} ({username})")

async def main():
    checker = TelegramChecker()
    await checker.initialize()
    
    while True:
        rprint("\n[bold cyan]Telegram Account Checker[/bold cyan]")
        rprint("\n1. Check phone numbers from input")
        rprint("2. Check phone numbers from Excel/Text file")
        rprint("3. Check usernames from input")
        rprint("4. Check usernames from text file")
        rprint("5. Clear saved credentials")
        rprint("6. Exit")
        
        choice = Prompt.ask("\nSelect an option", choices=["1", "2", "3", "4", "5", "6"])
        
        if choice == "1":
            phones = [p.strip() for p in Prompt.ask("Enter phone numbers (comma-separated)").split(",")]
            results = await checker.process_phones(phones)
        
        elif choice == "2":
            file_path = Prompt.ask("Enter the path to your Excel file (e.g., bd_numbers_state.xlsx)")
            try:
                # Check file extension to determine how to read it
                if file_path.lower().endswith(('.xlsx', '.xls')):
                    phones = read_excel_phones(file_path)
                else:
                    # Fallback to text file reading
                    with open(file_path, 'r') as f:
                        phones = [line.strip() for line in f if line.strip()]
                
                if not phones:
                    console.print("[red]No phone numbers found in the file![/red]")
                    continue
                    
                console.print(f"[green]Found {len(phones)} phone numbers in the file[/green]")
                results = await checker.process_phones(phones)
            except FileNotFoundError:
                console.print("[red]File not found![/red]")
                continue
            except Exception as e:
                console.print(f"[red]Error reading file: {str(e)}[/red]")
                continue
        
        elif choice == "3":
            usernames = [u.strip() for u in Prompt.ask("Enter usernames (comma-separated)").split(",")]
            results = await checker.process_usernames(usernames)
        
        elif choice == "4":
            file_path = Prompt.ask("Enter the path to your usernames text file")
            try:
                with open(file_path, 'r') as f:
                    usernames = [line.strip() for line in f if line.strip()]
                results = await checker.process_usernames(usernames)
            except FileNotFoundError:
                console.print("[red]File not found![/red]")
                continue
        
        elif choice == "5":
            if Confirm.ask("Are you sure you want to clear saved credentials?"):
                if CONFIG_FILE.exists(): CONFIG_FILE.unlink()
                if Path('telegram_checker_session.session').exists(): Path('telegram_checker_session.session').unlink()
                console.print("[green]Credentials cleared. Please restart the program.[/green]")
                break
            continue
        
        else:
            break
            
        if 'results' in locals():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Save to JSON
            json_file = RESULTS_DIR / f"results_{timestamp}.json"
            with open(json_file, 'w') as f:
                json.dump(results, f, indent=2)
            console.print(f"[green]Results saved to JSON: {json_file}[/green]")
            
            # Save to Excel
            excel_file = RESULTS_DIR / f"results_{timestamp}.xlsx"
            save_to_excel(results, excel_file)
            
            # Display summary
            display_summary(results)

if __name__ == "__main__":
    # Install required packages if not already installed
    try:
        import pandas
        import openpyxl  # For .xlsx files
    except ImportError:
        console.print("[yellow]Installing required packages for Excel support...[/yellow]")
        import subprocess
        subprocess.check_call(["pip", "install", "pandas", "openpyxl", "xlrd"])
        
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        console.print("\n[yellow]Program terminated by user[/yellow]")
    except Exception as e:
        console.print(f"\n[red]An error occurred: {str(e)}[/red]")
        logger.exception("Unhandled exception occurred")